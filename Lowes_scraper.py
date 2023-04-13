import re
import json
import csv
from time import sleep
import os
import warnings
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook


options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_argument("--headless=new")
options.add_experimental_option("excludeSwitches",["enable-automation"])
warnings.simplefilter(action='ignore', category=FutureWarning)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)
driver.get("https://www.lowes.com/")
sleep(1)
driver.add_cookie({'name': 'zipcode', 'value': '11763'})
sleep(2)
driver.refresh()

products=[]

def read_input(file_name):
    rows_data = []
    workbook = load_workbook(file_name)
    worksheet = workbook.active
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    for row in range(1, max_row+1):
        col_data = []
        for col in range(1, max_col+1):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                col_data.append(cell.value)
                rows_data.append(col_data)
            else:
                print(f"\n\t\tEmpty cell Founded row no: {row} col no: {col}\n")
    return rows_data


def save_as_xlsx(products):
    with open("output.csv", "w", newline='') as output_csv:
        writer = csv.writer(output_csv)
        writer.writerow(["productUrl","brand","title", "upc", "itemNumber", "modelNumber", "price($)", "availableForDelivery", "availableQuantityForDelivery", "freeShipping", "costOfDelivery","deliveryDate"])
        for product in products:
            writer.writerow(product.values())
    df = pd.read_csv('output.csv')
    convert_to_xlsx = pd.ExcelWriter('output.xlsx', engine='xlsxwriter')
    df.to_excel(convert_to_xlsx, sheet_name='Sheet1', index=False)
    convert_to_xlsx.save()
    os.system("del output.csv")


def get_product_links():
    sleep(1)
    product_links = driver.find_elements(By.XPATH, '//section[@id="listItems"]//div[contains(@class, "description-section")]//div[@data-selector="prd-description-holder"]//a')
    if len(product_links) == 0:
            product_links = driver.find_elements(By.XPATH, '//section[@id="listItems"]//div[@data-selector="prd-description-holder"]//a')
    return [product_link.get_attribute('href') for product_link in product_links]

def get_single_product_details(product_link):
    driver.delete_cookie("ak_bmsc")
    driver.get(product_link)
    sleep(3)
    data = re.search(r"window\['__PRELOADED_STATE__'] = (\{.*?})<", driver.page_source)
    data = json.loads(data.group(1))
    product_id_from_json = data["productId"]
    product_details_from_json = data["productDetails"][product_id_from_json]

    product_data = product_details_from_json["product"]
    location_data = product_details_from_json["location"]
    inventory_data = product_details_from_json["itemInventory"]


    try:
        title = product_data["title"]
    except:
        title = "N/A"
    try:
        brand = product_data["brand"]
    except:
        brand = "N/A"

    try:
        upc = product_data["barcode"]
    except:
        upc = "N/A"

    try:
        item_number = product_data["itemNumber"]
    except:
        item_number = "N/A"

    try:
        model_number = product_data["modelId"]
    except:
        model_number = "N/A"

    try:
        price = location_data["price"]["pricingDataList"][0]["finalPrice"]
    except:
        price = "N/A"

    try:
        free_shipping = location_data["promotion"]["freeDelivery"]
    except:
        free_shipping = "N/A"

    try:
        available_for_delivery = inventory_data["analyticsData"]["parcel"]["availabilityStatus"]
    except:
        available_for_delivery = "N/A"

    try:
        available_quantity_for_delivery = inventory_data["analyticsData"]["parcel"]["availableQuantity"]
    except:
        available_quantity_for_delivery = "N/A"

    cost_of_delivery = "N/A"


    # if not free_shipping:
    sleep(1)
    cost = driver.execute_script(
        """return document.evaluate('(//div[contains(@class, "fulfilment-messages")])[2]/div[1]', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue?.innerText""")

    delivery_date = "N/A"
    if cost is not None:
        parts = cost.split(':')
        delivery_date = parts[0]
        if len(parts) == 2:
            cost_of_delivery = parts[1].replace('From', '').strip()


    product_details = {
        "product url":product_link,
        "brand": brand,
        "title": title,
        "upc": upc,
        "itemNumber": item_number,
        "modelNumber": model_number,
        "price": price,
        "availableForDelivery": available_for_delivery,
        "availableQuantityForDelivery": available_quantity_for_delivery,
        "freeShipping": free_shipping,
        "costOfDelivery": cost_of_delivery,
        "deliveryDate":delivery_date
    }

    return product_details


def listTostring(s):
    string = ""
    for element in s:
        string+=element
    return string



def get_product_links_and_details(page_link,search_term):
    sleep(3)
    offset=24
    current_page_link=""
    length = ""
    driver.get(page_link)
    i = 1
    while True:
                if current_page_link == driver.current_url:
                    offset = 24
                    print("\t\tAll products scrapped\n")
                    break
                links = get_product_links()
                length = len(links)
                if length == 0:
                    print("Zero products in current page\n")
                    break

                current_page_link=driver.current_url
                print(f"\t\tTotal products in current page {length}")
                print(f"Product page {i}  Link :- {current_page_link}")
                for single_link in links:
                    product = get_single_product_details(single_link)
                    products.append(product)
                    length -=1
                    print(f"Remaining products {length}")
                i += 1    
                save_as_xlsx(products)
                print("Product page changed\n")
                sleep(1)
                if search_term == True:
                    next_link = page_link + "&offset=" + str(offset)
                else:
                    next_link = page_link + "?offset=" + str(offset)
                sleep(2)
                driver.get(next_link)
                sleep(2)
                offset += 24

def get_brands_link(brand_page_link):
    driver.get(brand_page_link)
    all_brands_link = driver.find_elements(By.CSS_SELECTOR, '#app > div:nth-child(2) > div > div > div:nth-child(12) > div > div > div > div > div > a ')
    if len(all_brands_link) == 0:
            all_brands_link = driver.find_elements(By.CSS_SELECTOR, '#app > div:nth-child(2) > div > div > div:nth-child(9) > div > div > div > div > div > a')     
    return [brand_link.get_attribute('href') for brand_link in all_brands_link]

def main():
    search_terms = read_input("input.xlsx")
    search_term = False
    for link in search_terms:
        temp_link = listTostring(link)
        print("Input query from input.xlsx :-",temp_link)
        if "/c/" in temp_link:
            brands_link = get_brands_link(temp_link)
            remaing_brands = len(brands_link)
            print(f"\t\tThere are total {remaing_brands} brands")
            if remaing_brands != 0: 
                for single_link in brands_link:
                    print(f"Current brand no  {remaing_brands}\nBrand link:- {single_link}")
                    search_term = True
                    get_product_links_and_details(single_link,search_term) 
                    remaing_brands -= 1

        elif "/pl/" in temp_link:
            get_product_links_and_details(temp_link,search_term)
        elif "/pd/"in temp_link:
            product_details = get_single_product_details(temp_link)
            products.append(product_details)
            save_as_xlsx(products)
        elif "/search?" in temp_link:
            search_term = True
            get_product_links_and_details(temp_link,search_term)
        else:
            search_term = True
            search_link = "https://www.lowes.com/search?searchTerm="+temp_link
            get_product_links_and_details(search_link,search_term)
        print("Input Query changed\n\n")


    
if __name__ == "__main__":
    main()
